#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PIF 難民特賣會 - Excel → 網站 同步腳本

功能：
  讀取最新的 Excel 檔，自動更新網站的：
    - 商品清單 (products.js)
    - 商品圖片 (images/)
    - 訂購須知 (info.json)
    - 庫存警示 (依 I 欄庫存，≤20 自動加「僅剩」標記)

使用：
  1. 修改 Excel 檔（改價格、新增商品、貼新圖片、改庫存等）
  2. 存檔並關閉 Excel
  3. 雙擊 一鍵同步.bat
"""

import os
import re
import sys
import json
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent.absolute()
PARENT_DIR = SCRIPT_DIR.parent
EXCEL_CANDIDATES = [f for f in PARENT_DIR.glob("*.xlsx") if not f.name.startswith("~$")]

PRODUCT_START_ROW = 13
PRODUCT_NAME_COL = 4
PRODUCT_BARCODE_COL = 3
PRODUCT_PRICE_COL = 5
PRODUCT_BOX_COL = 6
PRODUCT_STOCK_COL = 9
STOCK_WARN_THRESHOLD = 20


def safe_int(v, default=0):
    if v is None or v == "":
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def main():
    print("=" * 60)
    print("  PIF 難民特賣會 - Excel 同步腳本")
    print("=" * 60)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("\n[X] 缺少 openpyxl 套件，請先在 cmd 執行：")
        print("   pip install openpyxl\n")
        input("按 Enter 結束...")
        sys.exit(1)

    if not EXCEL_CANDIDATES:
        print(f"\n[X] 在 {PARENT_DIR} 找不到任何 .xlsx 檔案")
        input("按 Enter 結束...")
        sys.exit(1)

    # 檢測 Excel 是否還開啟
    lock_files = list(PARENT_DIR.glob("~$*"))
    if lock_files:
        print("\n" + "!" * 60)
        print("  [警告] 偵測到 Excel 檔案還開啟中！")
        print("!" * 60)
        print(f"  找到鎖定檔：{lock_files[0].name}")
        print("\n  請先存檔並關閉 Excel 再執行此腳本。\n")
        choice = input("  仍要繼續嗎？(輸入 y 繼續，其他鍵離開): ").strip().lower()
        if choice != "y":
            input("按 Enter 結束...")
            sys.exit(0)

    # 選 Excel 檔
    if len(EXCEL_CANDIDATES) == 1:
        excel_path = EXCEL_CANDIDATES[0]
    else:
        print(f"\n找到 {len(EXCEL_CANDIDATES)} 個 Excel 檔，請選擇：")
        for i, f in enumerate(EXCEL_CANDIDATES, 1):
            print(f"  {i}. {f.name}")
        choice = input(f"\n輸入編號 [預設 1]: ").strip() or "1"
        try:
            excel_path = EXCEL_CANDIDATES[int(choice) - 1]
        except (ValueError, IndexError):
            excel_path = EXCEL_CANDIDATES[0]

    print(f"\n[i] 讀取 Excel：{excel_path.name}")

    wb = load_workbook(excel_path)
    sheet = wb.active

    # 訂購須知
    info = {}
    for row in sheet.iter_rows(min_row=1, max_row=11, values_only=True):
        if row[0] and row[1] and isinstance(row[0], str):
            key = row[0].strip().replace("　", "").replace(" ", "")
            value = str(row[1]).strip() if row[1] else ""
            if key in ["取貨地點", "運費說明", "價格說明", "訂購單位", "截單日期", "庫存說明", "備註"]:
                info[key] = value
    print(f"  [OK] 讀取訂購須知：{len(info)} 項")

    # 商品資料
    products = []
    row_to_no = {}
    row = PRODUCT_START_ROW
    while True:
        no = sheet.cell(row=row, column=1).value
        if not isinstance(no, int):
            break
        name_raw = sheet.cell(row=row, column=PRODUCT_NAME_COL).value
        barcode = sheet.cell(row=row, column=PRODUCT_BARCODE_COL).value
        price = sheet.cell(row=row, column=PRODUCT_PRICE_COL).value
        box_qty = sheet.cell(row=row, column=PRODUCT_BOX_COL).value
        stock_qty = sheet.cell(row=row, column=PRODUCT_STOCK_COL).value

        if not name_raw:
            row += 1
            continue

        name_full = str(name_raw)

        # 庫存警示：以 I 欄為準 (≤20 自動加警示，>20 不顯示)
        # 名稱中若有「⚠ 僅剩 X 個」會直接移除，避免舊資料干擾
        name_no_stock = re.sub(r"\s*⚠\s*僅剩\s*\d+\s*個\s*", "", name_full).strip()
        if isinstance(stock_qty, (int, float)) and stock_qty <= STOCK_WARN_THRESHOLD:
            stock_left = max(0, int(stock_qty))
        else:
            stock_left = None

        parts = name_no_stock.split("\n")
        main_name = parts[0].strip()
        spec = " ".join(p.strip() for p in parts[1:]) if len(parts) > 1 else ""

        products.append({
            "no": no,
            "name": main_name,
            "spec": spec,
            "barcode": str(barcode).strip() if barcode else "",
            "price": safe_int(price),
            "box_qty": safe_int(box_qty),
            "stock_left": stock_left
        })
        row_to_no[row] = no
        row += 1

    print(f"  [OK] 讀取商品資料：{len(products)} 項")

    # 處理圖片
    images_dir = SCRIPT_DIR / "images"
    images_dir.mkdir(exist_ok=True)
    old_images = set(f.name for f in images_dir.iterdir() if f.is_file())
    new_images = set()

    saved_count = 0
    for img in sheet._images:
        anchor_row = img.anchor._from.row + 1
        if anchor_row in row_to_no:
            no = row_to_no[anchor_row]
            data = img._data()
            if data[:3] == b"\xff\xd8\xff":
                ext = "jpg"
            elif data[:8] == b"\x89PNG\r\n\x1a\n":
                ext = "png"
            else:
                ext = "jpg"
            filename = f"product_{no:03d}.{ext}"
            for other_ext in ("jpg", "png"):
                old_path = images_dir / f"product_{no:03d}.{other_ext}"
                if old_path.exists() and old_path.name != filename:
                    try:
                        old_path.unlink()
                    except Exception:
                        pass
            (images_dir / filename).write_bytes(data)
            new_images.add(filename)
            saved_count += 1
            for p in products:
                if p["no"] == no:
                    p["filename"] = filename
                    break

    print(f"  [OK] 寫入商品圖片：{saved_count} 張")

    # 缺新圖時保留舊圖
    fallback_count = 0
    for p in products:
        if p.get("filename"):
            continue
        no = p["no"]
        for ext in ("jpg", "png"):
            old_filename = f"product_{no:03d}.{ext}"
            if (images_dir / old_filename).exists():
                p["filename"] = old_filename
                new_images.add(old_filename)
                fallback_count += 1
                break
    if fallback_count > 0:
        print(f"  [i] 保留舊圖片：{fallback_count} 張")

    # 安全檢查
    if saved_count == 0 and fallback_count == 0 and len(products) > 0:
        print("\n[X] 沒有抓到任何圖片！可能 Excel 還開啟著。")
        print("    保險起見不覆蓋 products.js")
        input("按 Enter 結束...")
        sys.exit(1)

    # 刪除多餘圖片
    expected = set(p.get("filename") for p in products if p.get("filename"))
    redundant = old_images - expected
    deleted = 0
    for f in redundant:
        if f.startswith("product_9999"):
            continue  # 保留贈品圖
        try:
            (images_dir / f).unlink()
            deleted += 1
        except Exception:
            pass
    if deleted > 0:
        print(f"  [OK] 清除多餘圖片：{deleted} 張")

    # 缺圖警告
    missing = [p["no"] for p in products if not p.get("filename")]
    if missing:
        print(f"  [警告] 下列商品沒有圖片：No.{missing}")

    # 產生 products.js
    js_lines = ["const PRODUCTS = ["]
    for p in products:
        if not p.get("filename"):
            continue
        name_esc = p["name"].replace("\\", "\\\\").replace('"', '\\"')
        spec_esc = p["spec"].replace("\\", "\\\\").replace('"', '\\"')
        line = f'  {{ no: {p["no"]}, name: "{name_esc}", spec: "{spec_esc}", '
        line += f'barcode: "{p["barcode"]}", price: {p["price"]}, boxQty: {p["box_qty"]}'
        if p["stock_left"] is not None:
            line += f', stockLeft: {p["stock_left"]}'
        line += f', img: "images/{p["filename"]}" }},'
        js_lines.append(line)

    # 附上滿額贈商品 (No.9999) 如果原本就存在
    js_path = SCRIPT_DIR / "products.js"
    if js_path.exists():
        old_content = js_path.read_text(encoding="utf-8")
        gift_match = re.search(r'  \{ no: 9999.*?\},', old_content)
        if gift_match:
            js_lines.append(gift_match.group(0))

    js_lines.append("];")
    js_path.write_text("\n".join(js_lines) + "\n", encoding="utf-8")
    print(f"  [OK] 更新 products.js")

    # info.json
    info_path = SCRIPT_DIR / "info.json"
    info_path.write_text(json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  [OK] 更新 info.json")

    # 統計
    warn_count = sum(1 for p in products if p["stock_left"] is not None)
    print(f"\n{'=' * 60}")
    print(f"  [完成] 同步成功！")
    print(f"{'=' * 60}")
    print(f"  - 商品總數：{len(products)} 項")
    print(f"  - 含庫存警示：{warn_count} 項")
    print(f"  - 圖片總數：{len(new_images)} 張")
    print(f"\n  下一步：直接打開 index.html 測試，或重新部署到 GitHub Pages / Netlify")
    print(f"\n  [注意] 訂購須知區塊寫在 index.html 內，若要修改取貨地點/運費等請告訴 Claude")

    input("\n按 Enter 關閉視窗...")


if __name__ == "__main__":
    main()
